#!/usr/bin/env python3
"""Convert Excel files to CSV for comparison"""

import csv
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

def xlsx_to_csv(xlsx_file, csv_file):
    """Convert Excel file to CSV"""
    wb = openpyxl.load_workbook(xlsx_file)
    
    # Get all sheets
    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        
        # Determine output filename
        if len(wb.sheetnames) > 1:
            # Multiple sheets - add sheet name/number to filename
            base_name = csv_file.rsplit('.csv', 1)[0]
            output_file = f"{base_name}_sheet{sheet_idx+1}_{sheet_name.replace(' ', '_')}.csv"
        else:
            # Single sheet - use filename as-is
            output_file = csv_file
        
        # Write to CSV
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        
        print(f"  Exported sheet '{sheet_name}' to {output_file}")
    
    return len(wb.sheetnames)

def main():
    files = [
        ('New_LAMP_628302.xlsx', 'New_LAMP_628302.csv'),
        ('New_LAMP_628269.xlsx', 'New_LAMP_628269.csv')
    ]
    
    for xlsx_file, csv_file in files:
        print(f"\nConverting {xlsx_file}...")
        try:
            num_sheets = xlsx_to_csv(xlsx_file, csv_file)
            print(f"  ✓ Converted {num_sheets} sheet(s)")
        except Exception as e:
            print(f"  ✗ Error: {e}")
            import traceback
            traceback.print_exc()

if __name__ == '__main__':
    main()
